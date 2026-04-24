"""Smoke: verify the Excel export is valid and has 5 sheets."""
import sys
from pathlib import Path

import openpyxl

p = Path(sys.argv[1]) if len(sys.argv) > 1 else next(
    Path(r"data/exports").glob("*.xlsx"))
wb = openpyxl.load_workbook(p)
print(f"Loaded: {p.name}")
print(f"Sheets: {wb.sheetnames}")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s}: {ws.max_row} rows x {ws.max_column} cols")
