"""Prospects table column layout matches ``prospects (1) template.xlsx`` (87 columns)."""

from __future__ import annotations

from pathlib import Path

from src import config


def test_prospects_table_is_87_columns() -> None:
    assert len(config.PROSPECTS_TABLE_COLUMNS) == 87


def test_overall_is_column_35_and_driving_layup_36() -> None:
    """1-based Excel columns 35–36 == indices 34–35."""
    assert config.PROSPECTS_TABLE_COLUMNS.index("overall_2k") == 34
    assert config.PROSPECTS_TABLE_COLUMNS.index("driving_layup_2k") == 35


def test_bio_leading_fields_and_length() -> None:
    """Bio block: position-led prefix + ranks/names through status (14 cols)."""
    head = config.PROSPECTS_TABLE_COLUMNS[:14]
    assert head[:5] == (
        "pos",
        "secondary_position",
        "age",
        "height_in",
        "weight_lbs",
    )
    assert head == tuple(config.PROSPECTS_BIO_COLUMNS)
    assert config.PROSPECTS_TABLE_COLUMNS[14:34] == tuple(config.STAT_COLUMNS)


def test_post_hook_fade_positions() -> None:
    t = config.PROSPECTS_TABLE_COLUMNS
    assert t.index("post_hook_2k") == 42
    assert t.index("post_fade_2k") == 43


def test_draw_foul_then_shot_iq_then_ball_handle() -> None:
    t = config.PROSPECTS_TABLE_COLUMNS
    assert t.index("shot_iq_2k") == t.index("draw_foul_2k") + 1
    assert t.index("ball_handle_2k") == t.index("shot_iq_2k") + 1


def test_intangibles_hustle_durability_tail() -> None:
    """Template: intangibles col 69, hustle 70, durability 71 (1-based Excel)."""
    t = config.PROSPECTS_TABLE_COLUMNS
    assert t.index("intangibles_2k") + 1 == 69
    assert t.index("hustle_2k") + 1 == 70
    assert t.index("durability_2k") + 1 == 71


def test_meta_full_name_through_column1() -> None:
    tail = config.PROSPECTS_TABLE_COLUMNS[71:87]
    assert tail == tuple(config.PROSPECTS_META_COLUMNS)


def test_rating_tuple_matches_workbook_block() -> None:
    """Indices 34–70 (0-based) of PROSPECTS_TABLE_COLUMNS == RATING_ATTRIBUTES."""
    slab = config.PROSPECTS_TABLE_COLUMNS[34:71]
    assert slab == tuple(config.RATING_ATTRIBUTES)


def _normalize_template_header(cell) -> str:
    if cell is None:
        return ""
    s = str(cell).strip()
    aliases = {
        "Last Name": "last_name",
        "First Name": "first_name",
    }
    return aliases.get(s, s)


def test_optional_download_template_ratings_meta_alignment() -> None:
    """Ratings + meta headers align with workbook (canonical bio omits ``height_ft``)."""
    import pytest

    candidates = [
        Path(r"c:\Users\ofirs\Downloads\prospects (1) template.xlsx"),
        Path(__file__).resolve().parents[1] / "data" / "prospects 2026 (1).xlsx",
        Path(r"c:\Users\ofirs\OneDrive\NBA2K25\prospects 2026 (1).xlsx"),
    ]
    p = next((c for c in candidates if c.is_file()), None)
    if p is None:
        pytest.skip("No prospects template workbook on disk for header check")
    import openpyxl

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        name = "prospects" if "prospects" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[name]
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    finally:
        wb.close()
    raw = [_normalize_template_header(c) for c in row]
    raw = [h for h in raw if h][:87]
    assert len(raw) >= 71
    ours_r = list(config.PROSPECTS_TABLE_COLUMNS[34:71])
    tmpl_r = raw[34:71]
    assert tmpl_r == ours_r
    ours_m = list(config.PROSPECTS_TABLE_COLUMNS[71:87])
    tmpl_m = raw[71:87]
    assert tmpl_m == ours_m
