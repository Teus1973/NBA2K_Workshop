"""
openpyxl-based Excel writer.

Builds a four-sheet workbook (Reference / Prospects / Logs / Formulas). The
Prospects sheet can color-code cells by provenance when ``slug`` is included in
the exported columns (download slice omits it, so coloring is skipped there).

The downloaded **Prospects** sheet carries **names, bio, height (in/ft), full
per-game stat band, then 2K ratings** in template rating order. Freeze panes **C2**.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from .. import audit, config
from ..logger import get_logger
from . import data_loader

log = get_logger("exporters.excel_writer")

# Prospects download: stable column order (internal dataframe keys). Friendly
# headers for a few leading physical columns (see :data:`_EXCEL_HEADER_DISPLAY`).
PROSPECTS_EXCEL_DOWNLOAD_COLUMNS: tuple[str, ...] = (
    "last_name",
    "first_name",
    "pos",
    "secondary_position",
    "age",
    "height_in",
    "height_ft",
    "weight_lbs",
) + config.STAT_COLUMNS + config.RATING_ATTRIBUTES

_EXCEL_HEADER_DISPLAY: dict[str, str] = {
    "pos": "Pos",
    "age": "Age",
    "height_in": "Height (in)",
    "height_ft": "Height (ft)",
    "weight_lbs": "Weight (lbs)",
}


# Color palette for the Prospects sheet provenance coding (Documents/PLAN.md sec 2.2).
COLOR_SCRAPED = PatternFill("solid", fgColor="C6EFCE")  # green
COLOR_COMBINE = PatternFill("solid", fgColor="BDD7EE")  # blue
COLOR_FORMULA = PatternFill("solid", fgColor="FFEB9C")  # yellow
COLOR_MANUAL = PatternFill("solid", fgColor="D9D9D9")   # grey
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")


PROVENANCE_FILLS: dict[str, PatternFill] = {
    "scraped": COLOR_SCRAPED,
    "combine": COLOR_COMBINE,
    "formula": COLOR_FORMULA,
    "manual": COLOR_MANUAL,
}


_LONG_TEXT_COLS: frozenset[str] = frozenset({
    "notes",
    "full_name",
    "school_or_team",
    "added_by",
    "scouting_ai_summary",
    "scouting_physical_text",
    "scouting_physical_json",
    "manual_override_json",
    "league_stats",
    "source",
    "updated_at",
    "updated_at_stats",
    "computed_at",
})


def _apply_prospect_column_widths(ws, df: pd.DataFrame) -> None:
    """Stable widths on Prospects — avoids megabyte-wide columns from long text cells."""
    if df.empty:
        return
    stat_set = frozenset(config.STAT_COLUMNS)
    for idx, name in enumerate(df.columns, start=1):
        letter = get_column_letter(idx)
        if name in _LONG_TEXT_COLS:
            w = 30.0
        elif name.endswith("_2k"):
            w = 10.5
        elif name in stat_set:
            w = 10.5 if name in ("fg_pct", "fg3_pct", "ft_pct") else 9.5
        elif name == "slug":
            w = 16.0
        elif name in ("season", "league", "status", "column1", "formula_version"):
            w = 12.0
        elif name in ("espn_rank", "other_rank"):
            w = 11.0
        else:
            w = 11.5
        ws.column_dimensions[letter].width = min(max(w, 8.0), 44.0)


def _apply_prospects_download_headers(ws) -> None:
    """Replace row-1 labels for ``pos`` / ``age`` / ``height_in`` (Excel-only)."""
    if ws.max_row < 1:
        return
    for cell in ws[1]:
        if cell.value is None:
            continue
        key = str(cell.value).strip()
        disp = _EXCEL_HEADER_DISPLAY.get(key)
        if disp is not None:
            cell.value = disp


def _write_df(
    ws,
    df: pd.DataFrame,
    *,
    freeze: str | None = "A2",
    autosize_columns: bool = True,
) -> None:
    if df is None or df.empty:
        ws.append(["(no data)"])
        return
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        ws.append(row)
        if i == 0:
            for cell in ws[ws.max_row]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
    if freeze:
        ws.freeze_panes = freeze
    if not autosize_columns:
        return
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0
                      for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 36)


def _apply_prospect_provenance(
    ws,
    df: pd.DataFrame,
    provenance_by_slug: Mapping[str, Mapping[str, str]],
) -> None:
    if df.empty or not provenance_by_slug:
        return
    col_index = {name: idx + 1 for idx, name in enumerate(df.columns)}
    slug_col = col_index.get("slug")
    if slug_col is None:
        return
    for i, slug in enumerate(df["slug"].tolist(), start=2):
        prov = provenance_by_slug.get(slug) or {}
        for attr, source in prov.items():
            if attr not in col_index:
                continue
            fill = PROVENANCE_FILLS.get(source.split("+")[0])
            if fill is None:
                continue
            ws.cell(row=i, column=col_index[attr]).fill = fill


# ---------------------------------------------------------------------------
def export_to_excel(
    out_path: Path | str,
    *,
    provenance_by_slug: Mapping[str, Mapping[str, str]] | None = None,
    season: str = config.CURRENT_SEASON,
    season_type: str = "Regular",
) -> Path:
    """Write the four-sheet Excel workbook and return the output path."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_ref = wb.active
    ws_ref.title = "Reference"
    ws_pro = wb.create_sheet("Prospects")
    ws_eur = wb.create_sheet("Europeans")
    ws_log = wb.create_sheet("Logs")
    ws_for = wb.create_sheet("Formulas")

    ref_df = data_loader.load_reference_df(season, season_type)
    pro_df_full = data_loader.load_prospects_df()
    pro_df = pro_df_full.reindex(columns=list(PROSPECTS_EXCEL_DOWNLOAD_COLUMNS))
    log_df = data_loader.load_audit_df(limit=5000)
    for_df = data_loader.load_formulas_df()

    _write_df(ws_ref, ref_df)
    _write_df(ws_pro, pro_df, freeze=None, autosize_columns=False)
    _apply_prospects_download_headers(ws_pro)
    _apply_prospect_provenance(ws_pro, pro_df, provenance_by_slug or {})
    ws_pro.freeze_panes = "C2"
    _apply_prospect_column_widths(ws_pro, pro_df)
    _write_df(ws_eur, pd.DataFrame({"note": ["Phase 2 / deferred"]}))
    _write_df(ws_log, log_df)
    _write_df(ws_for, for_df.drop(columns=["yaml_blob"], errors="ignore")
              if not for_df.empty else for_df)

    wb.save(str(out_path))

    audit.log_event(
        action="export_excel",
        entity_type="export",
        entity_slug=out_path.name,
        note=(f"reference={len(ref_df)} prospects={len(pro_df)} "
              f"logs={len(log_df)} formulas={len(for_df)}"),
    )
    log.info("Excel export -> %s (%d prospects, %d reference)",
             out_path, len(pro_df), len(ref_df))
    return out_path
